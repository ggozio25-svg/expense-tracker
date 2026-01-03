from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import io
from PIL import Image
import base64
from decimal import Decimal
import json
import requests

# Carica variabili ambiente
load_dotenv()

app = Flask(__name__, static_folder='static')
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key')
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_UPLOAD_SIZE', 10485760))
CORS(app)

# Configurazione Supabase
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
SUPABASE_SERVICE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

# Headers per API Supabase
def get_supabase_headers(use_service_key=False):
    key = SUPABASE_SERVICE_KEY if use_service_key else SUPABASE_KEY
    return {
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

# OCR disabilitato (Google Vision non incluso)
vision_client = None

# Helper function per conversione Decimal
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)

app.json_encoder = DecimalEncoder

# ============= ROUTE FRONTEND =============

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

# ============= API CATEGORIE =============

@app.route('/api/categorie', methods=['GET'])
def get_categorie():
    try:
        url = f"{SUPABASE_URL}/rest/v1/categorie"
        params = {'attiva': 'eq.true'}
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        response.raise_for_status()
        return jsonify(response.json()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categorie', methods=['POST'])
def create_categoria():
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/categorie"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= API CLIENTI =============

@app.route('/api/clienti', methods=['GET'])
def get_clienti():
    try:
        url = f"{SUPABASE_URL}/rest/v1/clienti"
        params = {'attivo': 'eq.true', 'order': 'nome.asc'}
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        response.raise_for_status()
        return jsonify(response.json()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienti', methods=['POST'])
def create_cliente():
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/clienti"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienti/<int:id>', methods=['PUT'])
def update_cliente(id):
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/clienti"
        params = {'id': f'eq.{id}'}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienti/<int:id>', methods=['DELETE'])
def delete_cliente(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/clienti"
        params = {'id': f'eq.{id}'}
        data = {'attivo': False}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify({'message': 'Cliente disattivato'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= API PROGETTI =============

@app.route('/api/progetti', methods=['GET'])
def get_progetti():
    try:
        url = f"{SUPABASE_URL}/rest/v1/progetti"
        params = {'select': '*,clienti(nome)', 'order': 'data_inizio.desc'}
        cliente_id = request.args.get('cliente_id')
        if cliente_id:
            params['cliente_id'] = f'eq.{cliente_id}'
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        response.raise_for_status()
        return jsonify(response.json()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/progetti', methods=['POST'])
def create_progetto():
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/progetti"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/progetti/<int:id>', methods=['PUT'])
def update_progetto(id):
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/progetti"
        params = {'id': f'eq.{id}'}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= API VEICOLI =============

@app.route('/api/veicoli', methods=['POST'])
def create_veicolo():
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/veicoli"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        
        # Debug: mostra errore dettagliato
        if not response.ok:
            return jsonify({'error': f'{response.status_code} Client Error: {response.text}'}), 500
            
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/veicoli/<int:id>', methods=['PUT'])
def update_veicolo(id):
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/veicoli"
        params = {'id': f'eq.{id}'}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/veicoli/<int:id>', methods=['DELETE'])
def delete_veicolo(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/veicoli"
        params = {'id': f'eq.{id}'}
        data = {'attivo': False}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify({'message': 'Veicolo disattivato'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= API SPESE =============

@app.route('/api/spese', methods=['GET'])
def get_spese():
    try:
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {
            'select': '*,categorie(nome,colore),clienti(nome),progetti(nome)',
            'order': 'data_spesa.desc'
        }
        
        # Filtri opzionali
        if request.args.get('data_inizio'):
            params['data_spesa'] = f'gte.{request.args.get("data_inizio")}'
        if request.args.get('data_fine'):
            params['data_spesa'] = f'lte.{request.args.get("data_fine")}'
        if request.args.get('categoria_id'):
            params['categoria_id'] = f'eq.{request.args.get("categoria_id")}'
        if request.args.get('cliente_id'):
            params['cliente_id'] = f'eq.{request.args.get("cliente_id")}'
        if request.args.get('addebitabile'):
            params['addebitabile'] = f'eq.{request.args.get("addebitabile")}'
            
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        response.raise_for_status()
        return jsonify(response.json()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spese', methods=['POST'])
def create_spesa():
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/spese"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spese/<int:id>', methods=['PUT'])
def update_spesa(id):
    try:
        data = request.get_json()
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {'id': f'eq.{id}'}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spese/<int:id>', methods=['DELETE'])
def delete_spesa(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {'id': f'eq.{id}'}
        response = requests.delete(url, headers=get_supabase_headers(True), params=params)
        response.raise_for_status()
        return jsonify({'message': 'Spesa eliminata'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= API CHILOMETRICHE =============

@app.route('/api/chilometriche', methods=['GET'])
def get_chilometriche():
    try:
        url = f"{SUPABASE_URL}/rest/v1/chilometriche"
        params = {
            'select': '*,veicoli(targa,modello),clienti(nome),progetti(nome)',
            'order': 'data_viaggio.desc'
        }
        
        # Filtri opzionali
        if request.args.get('data_inizio'):
            params['data_viaggio'] = f'gte.{request.args.get("data_inizio")}'
        if request.args.get('data_fine'):
            params['data_viaggio'] = f'lte.{request.args.get("data_fine")}'
        if request.args.get('veicolo_id'):
            params['veicolo_id'] = f'eq.{request.args.get("veicolo_id")}'
        if request.args.get('cliente_id'):
            params['cliente_id'] = f'eq.{request.args.get("cliente_id")}'
            
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        response.raise_for_status()
        return jsonify(response.json()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chilometriche', methods=['POST'])
def create_chilometrica():
    try:
        data = request.get_json()
        
        # Calcola rimborso automaticamente
        km = float(data.get('km_percorsi', 0))
        tariffa = float(data.get('tariffa_applicata', 0.19))
        data['rimborso_calcolato'] = round(km * tariffa, 2)
        
        url = f"{SUPABASE_URL}/rest/v1/chilometriche"
        response = requests.post(url, headers=get_supabase_headers(True), json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chilometriche/<int:id>', methods=['PUT'])
def update_chilometrica(id):
    try:
        data = request.get_json()
        
        # Ricalcola rimborso se necessario
        if 'km_percorsi' in data or 'tariffa_applicata' in data:
            km = float(data.get('km_percorsi', 0))
            tariffa = float(data.get('tariffa_applicata', 0.19))
            data['rimborso_calcolato'] = round(km * tariffa, 2)
        
        url = f"{SUPABASE_URL}/rest/v1/chilometriche"
        params = {'id': f'eq.{id}'}
        response = requests.patch(url, headers=get_supabase_headers(True), params=params, json=data)
        response.raise_for_status()
        return jsonify(response.json()[0]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chilometriche/<int:id>', methods=['DELETE'])
def delete_chilometrica(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/chilometriche"
        params = {'id': f'eq.{id}'}
        response = requests.delete(url, headers=get_supabase_headers(True), params=params)
        response.raise_for_status()
        return jsonify({'message': 'Chilometrica eliminata'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= UPLOAD IMMAGINE =============

@app.route('/api/upload', methods=['POST'])
def upload_image():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nessun file caricato'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome file vuoto'}), 400
        
        # Comprimi immagine
        img = Image.open(file)
        img.thumbnail((1200, 1200))
        
        # Converti in JPEG
        buffer = io.BytesIO()
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        img.save(buffer, format='JPEG', quality=85)
        buffer.seek(0)
        
        # Upload a Supabase Storage
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filename = filename.replace(' ', '_')
        
        storage_url = f"{SUPABASE_URL}/storage/v1/object/expenses/{filename}"
        headers = {
            'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
            'Content-Type': 'image/jpeg'
        }
        
        response = requests.post(storage_url, headers=headers, data=buffer.getvalue())
        
        if response.status_code not in [200, 201]:
            return jsonify({'error': 'Errore upload immagine'}), 500
        
        # URL pubblico dell'immagine
        image_url = f"{SUPABASE_URL}/storage/v1/object/public/expenses/{filename}"
        
        # OCR disabilitato
        ocr_data = {
            'message': 'OCR non disponibile - inserire dati manualmente'
        }
        
        return jsonify({
            'image_url': image_url,
            'ocr_data': ocr_data
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= STATISTICHE =============

@app.route('/api/stats/dashboard', methods=['GET'])
def get_dashboard_stats():
    try:
        oggi = datetime.now()
        primo_giorno_mese = oggi.replace(day=1).strftime('%Y-%m-%d')
        
        stats = {}
        
        # Spese mese corrente
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {'data_spesa': f'gte.{primo_giorno_mese}', 'select': 'importo'}
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        spese = response.json()
        stats['spese_mese'] = sum(float(s['importo']) for s in spese)
        
        # Spese addebitabili
        params = {'data_spesa': f'gte.{primo_giorno_mese}', 'addebitabile': 'eq.true', 'select': 'importo'}
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        spese_add = response.json()
        stats['spese_addebitabili'] = sum(float(s['importo']) for s in spese_add)
        
        # Km mese corrente
        url = f"{SUPABASE_URL}/rest/v1/chilometriche"
        params = {'data_viaggio': f'gte.{primo_giorno_mese}', 'select': 'km_percorsi,rimborso_calcolato'}
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        km = response.json()
        stats['km_mese'] = sum(float(k['km_percorsi']) for k in km)
        stats['rimborsi_km'] = sum(float(k['rimborso_calcolato']) for k in km)
        
        # Ultime spese
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {
            'select': '*,categorie(nome,colore),clienti(nome)',
            'order': 'data_spesa.desc',
            'limit': '10'
        }
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        stats['ultime_spese'] = response.json()
        
        # Spese per categoria (mese corrente)
        url = f"{SUPABASE_URL}/rest/v1/spese"
        params = {
            'data_spesa': f'gte.{primo_giorno_mese}',
            'select': 'importo,categorie(nome,colore)'
        }
        response = requests.get(url, headers=get_supabase_headers(), params=params)
        spese_cat = response.json()
        
        categorie_totali = {}
        for spesa in spese_cat:
            if spesa.get('categorie'):
                cat_nome = spesa['categorie']['nome']
                if cat_nome not in categorie_totali:
                    categorie_totali[cat_nome] = {
                        'nome': cat_nome,
                        'colore': spesa['categorie']['colore'],
                        'totale': 0
                    }
                categorie_totali[cat_nome]['totale'] += float(spesa['importo'])
        
        stats['spese_per_categoria'] = list(categorie_totali.values())
        
        return jsonify(stats), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= EXPORT EXCEL =============

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.get_json()
        tipo = data.get('tipo', 'spese')
        filtri = data.get('filtri', {})
        
        wb = Workbook()
        ws = wb.active
        
        if tipo == 'spese':
            ws.title = "Spese"
            headers = ['Data', 'Categoria', 'Cliente', 'Progetto', 'Descrizione', 'Importo', 'Fornitore', 'Addebitabile']
            ws.append(headers)
            
            # Formattazione header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Recupera spese
            url = f"{SUPABASE_URL}/rest/v1/spese"
            params = {
                'select': '*,categorie(nome),clienti(nome),progetti(nome)',
                'order': 'data_spesa.desc'
            }
            if filtri.get('data_inizio'):
                params['data_spesa'] = f'gte.{filtri["data_inizio"]}'
            if filtri.get('data_fine'):
                params['data_spesa'] = f'lte.{filtri["data_fine"]}'
            if filtri.get('cliente_id'):
                params['cliente_id'] = f'eq.{filtri["cliente_id"]}'
                
            response = requests.get(url, headers=get_supabase_headers(), params=params)
            spese = response.json()
            
            for spesa in spese:
                ws.append([
                    spesa['data_spesa'],
                    spesa['categorie']['nome'] if spesa.get('categorie') else '',
                    spesa['clienti']['nome'] if spesa.get('clienti') else '',
                    spesa['progetti']['nome'] if spesa.get('progetti') else '',
                    spesa['descrizione'],
                    float(spesa['importo']),
                    spesa.get('fornitore', ''),
                    'Sì' if spesa['addebitabile'] else 'No'
                ])
        
        elif tipo == 'chilometriche':
            ws.title = "Chilometriche"
            headers = ['Data', 'Veicolo', 'Partenza', 'Arrivo', 'Km', 'Tariffa', 'Rimborso', 'Cliente', 'Addebitabile']
            ws.append(headers)
            
            # Formattazione header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Recupera chilometriche
            url = f"{SUPABASE_URL}/rest/v1/chilometriche"
            params = {
                'select': '*,veicoli(targa),clienti(nome)',
                'order': 'data_viaggio.desc'
            }
            if filtri.get('data_inizio'):
                params['data_viaggio'] = f'gte.{filtri["data_inizio"]}'
            if filtri.get('data_fine'):
                params['data_viaggio'] = f'lte.{filtri["data_fine"]}'
            if filtri.get('veicolo_id'):
                params['veicolo_id'] = f'eq.{filtri["veicolo_id"]}'
                
            response = requests.get(url, headers=get_supabase_headers(), params=params)
            chilometriche = response.json()
            
            for km in chilometriche:
                ws.append([
                    km['data_viaggio'],
                    km['veicoli']['targa'] if km.get('veicoli') else '',
                    km['partenza'],
                    km['arrivo'],
                    float(km['km_percorsi']),
                    float(km['tariffa_applicata']),
                    float(km['rimborso_calcolato']),
                    km['clienti']['nome'] if km.get('clienti') else '',
                    'Sì' if km['addebitabile'] else 'No'
                ])
        
        # Salva file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= HEALTH CHECK =============

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'supabase_configured': bool(SUPABASE_URL and SUPABASE_KEY),
        'vision_configured': vision_client is not None
    }), 200

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
